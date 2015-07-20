#!/usr/bin/env python

"""
Python rewrite of http: //www.vlfeat.org/applications/caltech-101-code.html
"""

from os.path import exists, isdir, basename, join, splitext, isfile
from os import makedirs, remove, mkdir, listdir
from glob import glob
from random import sample, seed
from scipy import ones, mod, arange, array, where, ndarray, hstack, linspace, histogram, vstack, amax, amin
from scipy.misc import imread, imresize, imsave
from scipy.cluster.vq import vq
import numpy
import numpy as np
from vl_phow import vl_phow
from vlfeat import vl_ikmeans
from scipy.io import loadmat, savemat
from sklearn import svm
from sklearn.metrics import confusion_matrix, accuracy_score, classification_report, precision_score
import matplotlib.pyplot as plt
from datetime import datetime
from sklearn.kernel_approximation import AdditiveChi2Sampler
from cPickle import dump, load
import argparse
import multiprocessing
import sys
from openpyxl import *
import ftplib
from skimage import color
from PIL import Image, ImageChops
from scipy import ndimage
from scipy.ndimage import interpolation

IDENTIFIER = '2014-04-17-UR'
PREFIX = 'baseline'

FEATUREMAP = True
OVERWRITE = True  # DON'T load mat files generated with a different seed!!!
SAMPLE_SEED = 1963543398
VERBOSE = True	# set to 'SVM' if you want to get the svm output

class Configuration(object):
	def __init__(self, identifier='', prefix=''):
		self.calDir = '../../../data/2014_winter/256x256/vlfeat_training_jpg'

		# Path where training data will be stored
		self.dataDir = '../result-45-90'	 # should be resultDir or so
		if not exists(self.dataDir):
			makedirs(self.dataDir)
			print ("folder " + self.dataDir + " created")

		# Sum of these two numbers should be <= # of images in smallest
		# class
		self.numTrain = 70
		self.numTest = 30
		
		self.numCore = multiprocessing.cpu_count()
		self.imagesperclass = self.numTrain + self.numTest
		self.numClasses = 10
		self.numWords = 600
		self.numSpatialX = [2, 4]
		self.numSpatialY = [2, 4]
		self.quantizer = 'vq'  # kdtree from the .m version not implemented
		self.svmC = 15
		self.phowOpts = PHOWOptions(Verbose=False, Sizes=[2, 4, 6, 8], Step=3)
		self.prefix = prefix
		self.verbose = True
		self.extensions = [".jpg", ".jpeg", ".bmp", ".png", ".pgm", ".tif", ".tiff"]
		self.images_for_histogram = 30
		self.numbers_of_features_for_histogram = 100000
		self.imSize = 480
		
		self.saveFig = False
		self.showFig = False
		self.removeBg = True
		self.crop = True
		self.augment = True
		self.threshold = 1.05
		self.images = {}
		self.rotation = [-90, -45, 45, 90]
		
		self.vocabPath = join(self.dataDir, self.prefix + '-' + identifier + '-vocab.py.mat')
		self.histPath = join(self.dataDir, self.prefix + '-'  + identifier + '-hists.py.mat')
		self.modelPath = join(self.dataDir, self.prefix + '-' + identifier + '-model.py.mat')
		self.resultPath = join(self.dataDir, self.prefix + '-' + identifier + '-result')
		self.imageCropPath = join(self.dataDir, 'images/')

		# tests and conversions
		self.phowOpts.Sizes = ensure_type_array(self.phowOpts.Sizes)
		self.numSpatialX = ensure_type_array(self.numSpatialX)
		self.numSpatialY = ensure_type_array(self.numSpatialY)
		if (self.numSpatialX != self.numSpatialY).any():
			messageformat = [str(self.numSpatialX), str(self.numSpatialY)]
			message = "(self.numSpatialX != self.numSpatialY), because {0} != {1}".format(*messageformat)
			raise ValueError(message)

	def setClasses(self, classes):
		self.classes = classes #Save class names for later use by birdid_classifier

class Model(object):
	def __init__(self, classes, conf, vocab=None):
		self.classes = classes
		self.phowOpts = conf.phowOpts
		self.numSpatialX = conf.numSpatialX
		self.numSpatialY = conf.numSpatialY
		self.quantizer = conf.quantizer
		self.vocab = vocab

class PHOWOptions(object):
	def __init__(self, Verbose, Sizes, Step):
		self.Verbose = Verbose
		self.Sizes = Sizes
		self.Step = Step

def ensure_type_array(data):
	if (type(data) is not ndarray):
		if (type(data) is list):
			data = array(data)
		else:
			data = array([data])
	return data

def standardizeImage(im): #Scales image down to 640x480 or whatever the correct aspect ratio is with conf.imSize as the height
	im = array(im, 'float32') 
	if im.shape[0] > conf.imSize:
		resize_factor = float(conf.imSize) / im.shape[0]# don't remove trailing .0 to avoid integer devision
		im = imresize(im, resize_factor)
	if amax(im) > 1.1:
		im = im / 255.0
	assert((amax(im) > 0.01) & (amax(im) <= 1))
	assert((amin(im) >= 0.00))
	return im

def getPhowFeatures(imagedata, phowOpts): #extracts features from image
	im = standardizeImage(imagedata) #scale image to 640x480
	frames, descrs = vl_phow(im, verbose=phowOpts.Verbose, sizes=phowOpts.Sizes, step=phowOpts.Step)
	return frames, descrs

def getPhowFeaturesMulti(imagedata, phowOpts, idx): #used in multiprocessing for training vocab
	return [idx, getPhowFeatures(imagedata, phowOpts)[1]]

def getImageDescriptor(model, im, idx): #gets histograms
	im = standardizeImage(im) #scale image to 640x480
	height, width = im.shape[:2]
	numWords = model.vocab.shape[1]
	frames, descrs = getPhowFeatures(im, conf.phowOpts) #extract features
	# quantize appearance
	if model.quantizer == 'vq':
		binsa, _ = vq(descrs.T, model.vocab.T) #slowest function - assigns words from vocab to features in descrs
	elif model.quantizer == 'kdtree':
		raise ValueError('quantizer kdtree not implemented')
	else:
		raise ValueError('quantizer {0} not known or understood'.format(model.quantizer))
	hist = []
	#generate the histogram bins
	for n_spatial_bins_x, n_spatial_bins_y in zip(model.numSpatialX, model.numSpatialX):
		binsx, distsx = vq(frames[0, :], linspace(0, width, n_spatial_bins_x))
		binsy, distsy = vq(frames[1, :], linspace(0, height, n_spatial_bins_y))
		# binsx and binsy list to what spatial bin each feature point belongs to
		if (numpy.any(distsx < 0)) | (numpy.any(distsx > (width/n_spatial_bins_x+0.5))):
			print ("something went wrong")
			import pdb; pdb.set_trace()
		if (numpy.any(distsy < 0)) | (numpy.any(distsy > (height/n_spatial_bins_y+0.5))):
			print ("something went wrong")
			import pdb; pdb.set_trace()
		# combined quantization
		number_of_bins = n_spatial_bins_x * n_spatial_bins_y * numWords
		temp = arange(number_of_bins)
		# update using this: http://stackoverflow.com/questions/15230179/how-to-get-the-linear-index-for-a-numpy-array-sub2ind
		temp = temp.reshape([n_spatial_bins_x, n_spatial_bins_y, numWords])
		bin_comb = temp[binsx, binsy, binsa]
		hist_temp, _ = histogram(bin_comb, bins=range(number_of_bins+1), density=True) #generate histogram
		hist.append(hist_temp)

	hist = hstack(hist)
	hist = array(hist, 'float32') / sum(hist)
	numTot = float(conf.numClasses*(conf.numTrain+conf.numTest)*(len(conf.rotation)+1))
	sys.stdout.write ("\r"+str(datetime.now())+" Histograms Calculated: "+str(((idx+1)/numTot)*100.0)[:5]+"%") #make progress percentage
	sys.stdout.flush()
	return [idx, hist]

def get_classes(datasetpath, numClasses): #find classes in the data folder
	classes_paths = [files
					 for files in glob(datasetpath + "/*")
					 if isdir(files)]
	classes_paths.sort()
	classes = [basename(class_path) for class_path in classes_paths]
	if len(classes) == 0:
	   raise ValueError('no classes found')
	if len(classes) < numClasses:
	   raise ValueError('conf.numClasses is bigger than the number of folders')
	classes = classes[:numClasses]
	return classes

def get_imgfiles(path, extensions): #get images from 1 class folder
	all_files = []
	all_files.extend([join(path, basename(fname))
					 for fname in glob(path + "/*")
					 if splitext(fname)[-1].lower() in extensions])
	return all_files

def get_all_images(classes, conf): #gets all images from all classes
	all_images = []
	all_images_class_labels = []
	images_per_class = [0]
	for i, imageclass in enumerate(classes):
		path = join(conf.calDir, imageclass)
		extensions = conf.extensions
		imgs = get_imgfiles(path, extensions)
		if len(imgs) == 0:
			raise ValueError('no images for class ' + str(imageclass))
		all_images = all_images + imgs
		if len(images_per_class)>0:
			sum = images_per_class[len(images_per_class)-1]
		images_per_class.append(sum+len(imgs))
		class_labels = list(i * ones(len(imgs)))
		all_images_class_labels = all_images_class_labels + class_labels

	all_images_class_labels = array(all_images_class_labels, 'int')
	return all_images, all_images_class_labels, images_per_class

def create_split(all_images, images_per_class, conf): #split files between training and testing
	train_test = []
	for i in range(0, conf.numClasses):
		new_train = sample(all_images[images_per_class[i]:images_per_class[i+1]], conf.imagesperclass)
		train_test = train_test+new_train
	selTrain = []
	for i in range(0, conf.numClasses):
		selTrain = selTrain + sample(train_test[i*(conf.imagesperclass):(i+1)*(conf.imagesperclass)], conf.numTrain)
	selTest = [x for x in train_test if x not in selTrain]
	test = []
	train = []
	for i, ii in enumerate(selTrain):
		train.append(ii)
		selTrain[i] = all_images.index(ii)
	for i, ii in enumerate(selTest):
		test.append(ii)
		selTest[i] = all_images.index(ii)

	#crop all images
	if conf.crop:
		if not isdir(conf.imageCropPath):
			mkdir(conf.imageCropPath)
		pool = multiprocessing.Pool(processes=multiprocessing.cpu_count())
		result = [pool.apply_async(autoCrop, args=(i, img)) for i, img in enumerate(train+test)]
		res = [p.get() for p in result]
		pool.terminate()
		for r in res:
			conf.images[r[0]] = r[1]
		print ""
		print str(datetime.now())+" Done crop"
	else:
		for img in train:
			conf.images[img] = [img]
		for img in test:
			conf.images[img] = [img]
	
	return selTrain, selTest

def trim(im, color): #crop based on the binary image to zoom into the largest area
    bg = Image.new(im.mode, im.size, 0)
    diff = ImageChops.difference(im, bg)
    diff = ImageChops.add(diff, diff, 2.0, -100)
    bbox = diff.getbbox()
    if bbox:
        return color.crop(bbox) #actually returns the cropped image color, not im

def autoCrop(imName, img): #background remove and then crop
	extension = -(len(img.rpartition('.')[2])+1) #find how long the extension is, ie .jpg
	imageName = img.rpartition('/')[2][:extension] #get just the image name minus the extension and path
	if isfile(conf.imageCropPath+str(imageName)+"_0.jpg"):
		imAug = [conf.imageCropPath+str(imageName)+"_0.jpg"]
		if conf.augment:
			allRotated = True
			for rot in conf.rotation:
				if isfile(conf.imageCropPath+str(imageName)+"_"+str(rot)+".jpg"):
					imAug.append(conf.imageCropPath+str(imageName)+"_"+str(rot)+".jpg")
				else:
					allRotated = False
			if allRotated:
				return [img, imAug]
		else:
			return [img, imAug]
	im = imread(img)
	imOrig = imread(img)
	imAug = [conf.imageCropPath+str(imageName)+"_0.jpg"]
	x, y, z = im.shape
	binary_im = np.empty([x,y],np.uint8)
	r,g,b=Image.fromarray(im).getpixel((0,0))
	for i in range(0,x):
		for j in range(0,y):
			if im[i,j,1] > im[i,j,0]*conf.threshold and im[i,j,1] > im[i,j,2]*conf.threshold:
				im[i,j,:] = 255
				binary_im[i,j] = 0
			else:
				binary_im[i,j] = 1
	labels, numL = ndimage.label(binary_im) #find regions
	sizes = ndimage.sum(binary_im,labels,range(1,numL+1)) #find sizes of regions
	map = np.where(sizes==sizes.max())[0] + 1 #find largest region
	max_index = np.zeros(numL + 1, np.uint8)
	max_index[map] = 255
	max_feature = max_index[labels]
	
	if not conf.removeBg:
		imCrop = imOrig
	else:
		imCrop = im
	imCrop = trim(Image.fromarray(max_feature), Image.fromarray(imCrop)) #crop image
	x,y = imCrop.size #get cropped image size

	if x*y>1200000: #if cropped image is too large, then try again using HSV for bg removal
		x, y, z = imOrig.shape
		im2 = imOrig
		im2 = np.array(im2, np.float64)
		binary_im2 = np.empty([x,y], np.uint8)
		for i in range(0,x):
			for j in range(0,y):
				for k in range(0,z):
					im2[i,j,k] = im2[i,j,k]/255.0 #convert RGB from 0-255 to 0-1
		im2 = color.rgb2hsv(im2) #convert RGB to HSV
		for i in range(0,x):
			for j in range(0,y):
				for k in range(0,z):
					if k==0:
						v = 360 #change H from 0-1 to 0-360
					else:
						v = 100 #change S and V from 0-1 to 0-100
					im2[i,j,k] = int(im2[i,j,k]*v)
				if im2[i,j,0]<200 and im2[i,j,0]>60 and im2[i,j,1]>10 and im2[i,j,2]>10: #check H, S, and V if they are within the threshold
					binary_im2[i,j] = 0
				else:
					binary_im2[i,j] = 1

		labels, numL = ndimage.label(binary_im2) #find regions
		sizes = ndimage.sum(binary_im,labels,range(1,numL+1)) #find sizes of regions
		map = np.where(sizes==sizes.max())[0] + 1 #find largest region
		max_index = np.zeros(numL + 1, np.uint8)
		max_index[map] = 255
		max_feature = max_index[labels]

		if not conf.removeBg: #check if background should be included or removed in final output
			imCrop = imOrig
		else:
			imCrop = im
		imCrop = trim(Image.fromarray(max_feature), Image.fromarray(imCrop)) #crop image
		binary_im = binary_im2

	imsave(conf.imageCropPath+str(imageName)+"_0.jpg", imCrop) #save final photo

	if conf.augment:
		if not conf.removeBg: #check if background should be included or removed in final output
			imCrop = imOrig
		else:
			imCrop = im
		for rot in conf.rotation:
			imCropped = trim(Image.fromarray(interpolation.rotate(max_feature, rot, reshape=False)), Image.fromarray(interpolation.rotate(imCrop, rot, reshape=False)))
			imsave(conf.imageCropPath+str(imageName)+"_"+str(rot)+".jpg", imCropped)
			imAug.append(conf.imageCropPath+str(imageName)+"_"+str(rot)+".jpg")
	
	sys.stdout.write("\r"+str(datetime.now())+" AutoCropped Images: "+str((imName/float((conf.numTrain+conf.numTest)*conf.numClasses))*100.0)[:5]+"%") #make progress percentage
	sys.stdout.flush()
	return [img, imAug]

def trainVocab(selTrain, all_images, conf):
	selTrainFeats = sample(selTrain, conf.images_for_histogram)
	descrs = []
	imageFiles = []
	imgs = []
	for x in range(0,len(selTrainFeats)):
		imageFiles.append("hi")
	for i, ii in enumerate(selTrainFeats):
			imageFiles[i] = str(all_images[ii])
	for files in imageFiles:
			imgs.append(conf.images[files][0])
	#start multiprocessing block
	pool = multiprocessing.Pool(processes=conf.numCore)
	results = [pool.apply_async(getPhowFeaturesMulti, args=(imread(ii), conf.phowOpts, i)) for i, ii in enumerate(imgs)]
	descrs = [p.get() for p in results]
	pool.terminate()
	sorted(descrs)
	for descr in descrs:
		descr.pop(0)
	depict = []
	for descr in descrs:
		depict.append(descr[0])
	descrs = depict
	#end multiprocessing block
	descrs = hstack(descrs)
	n_features = descrs.shape[1]
	sample_indices = sample(arange(n_features), conf.numbers_of_features_for_histogram)
	descrs = descrs[:, sample_indices]
	descrs = array(descrs, 'uint8')
	
	# Quantize the descriptors to get the visual words
	vocab, _ = vl_ikmeans(descrs,
						  K=conf.numWords,
						  verbose=conf.verbose,
						  method='elkan')
	return vocab

def computeHistograms(selTrain, selTest, all_images, model, conf):
	imgs = []
	images = np.append(selTrain, selTest)
	imageFiles = []
	for x in range(0,len(images)):
		imageFiles.append("hi")
	for i, ii in enumerate(images):
			imageFiles[i] = str(all_images[ii])
	for files in imageFiles:
		for i in conf.images[files]:
			imgs.append(i)
	hists = []
	#start multiprocessing block
	pool = multiprocessing.Pool(processes=conf.numCore)
	results = [pool.apply_async(getImageDescriptor, args=(model, imread(str(im)), ii)) for ii, im in enumerate(imgs)]
	hists = [p.get() for p in results]
	pool.terminate()
	sorted(hists)
	for hist in hists:
		hist.pop(0)
	#end multiprocessing block
	hists = vstack(hists)
	print "" #puts in a new line to separate histogram percentage
	return hists

def saveCSV(file, accuracy, precision):
	dat = []
	dat.append(datetime.now())
	dat.append(str(PREFIX))
	dat.append(str(IDENTIFIER))
	dat.append(str(conf.phowOpts.Sizes))
	dat.append(str(SAMPLE_SEED))
	dat.append(str(accuracy))
	dat.append(str(conf.numTrain))
	dat.append(str(conf.numTest))
	dat.append(str(conf.numClasses))
	dat.append(str(conf.calDir))
	dat.append(str(conf.imSize))
	dat.append(str(conf.numWords))
	dat.append(str(conf.numbers_of_features_for_histogram))
	dat.append(str(conf.rotation))
	dat.append(str(precision))

	if isfile("phow_results.xlsx"): #create backup spreadsheet in case network is unmounted
		wb = load_workbook("phow_results.xlsx", guess_types=True)
		ws = wb.active
	else:
		wb = Workbook(guess_types=True)
		ws = wb.active
		ws.append(['Time Completed', 'Prefix', 'Identifier', 'Dsift Sizes', 'Sample Seed', 'Accuracy', 'Number of Train', 'Number of Test', 'Number of Classes', 'Image Path', 'Image Resize Height', 'Number of K-Means Centroids', 'Number of Histogram Features', 'Rotation', 'Precision'])
	ws.append(dat)
	wb.save("phow_results.xlsx")

	ftp = ftplib.FTP('dombrowskivpn.mynetgear.com', "lbarnett-students", 'lbarnett-studentaccess')#enter server information here
	ftp.set_pasv(False)
	with open("temp.xlsx", 'wb') as f:
		def callback(data):
			f.write(data)
		ftp.retrbinary('RETR '+str(file), callback)
	wb = load_workbook("temp.xlsx", guess_types=True)
	ws = wb.active
	ws.append(dat)
	wb.save("temp.xlsx")
	ftp.storbinary('STOR '+str(file), open('temp.xlsx','r'))
	ftp.close()
	remove('temp.xlsx')

def showFig(images, conf):
	axes = {}
	if len(images)>0:
		if len(images)<=16:
			x = 4
			y = 4
		elif len(images)<=25:
			x = 5
			y = 5
		elif len(images)<=35:
			x = 7
			y = 5
		else:
			x = 7
			y = 5
			while (int(x)*y)<len(images):
				y = y+1
				x = y*(1.75)
		x = int(x)
		fig = plt.figure(figsize=(16,10))
		for i, im in enumerate(images):
			axes[i] = fig.add_subplot(x,y,i+1)
			axes[i].imshow(imread(conf.images[im[0]][0]))
			axes[i].get_xaxis().set_ticks([])
			axes[i].get_yaxis().set_ticks([])
			axes[i].set_title("Classified as a "+conf.classes[im[1]['predictedclass']]+"\nActually is a "+conf.classes[im[1]['trueclass']])
	
	fig.set_tight_layout(True)
	plt.show()

def newAccuracy(true_classes, predicted_classes):
	misid = 0
	wrong = []
	newTrue = []
	newPrediction = []
	classGuess = np.zeros(len(conf.rotation)+1)
	for i in range(0, conf.numTest*conf.numClasses*(len(conf.rotation)+1)):
		classGuess[i%(len(conf.rotation)+1)] = predicted_classes[i]
		if i%(len(conf.rotation)+1) == len(conf.rotation):
			classGuess = list(classGuess)
			finalGuess = max(classGuess, key=classGuess.count)
			newTrue.append(true_classes[i])
			newPrediction.append(finalGuess)
			if finalGuess != true_classes[i]:
				misid = misid+1
				wrong.append([all_images[selTest[i]], {'trueclass':true_classes[i],'predictedclass':finalGuess}, i])
	originalNumBirds = float(len(true_classes)/(len(conf.rotation)+1))
	return [(1.0-(misid/originalNumBirds)), wrong, newTrue, newPrediction]

################
# Main Program #
################

if __name__ == '__main__':
	#################################
	# Handle command-line arguments #
	#################################
	parser = argparse.ArgumentParser()
	parser.add_argument("--sample_seed",
		help="Seed for choosing training sample", type=int)
	parser.add_argument("--identifier",
		help="Identifier for this data set. Should not contain the character '-'")
	parser.add_argument("--prefix",
		help="Tag used to distinguish versions of a data set")
	parser.add_argument("--image_dir",
						help="Path to directory containing images")
	parser.add_argument("--num_classes",
						help="Number of categories in image set",
						type=int)
	parser.add_argument("--num_train",
						help="Number of training images to use from each category",
						type=int)
	parser.add_argument("--num_test",
						help="Number of test images to use from each catetory",
						type=int)
	parser.add_argument("--dsift_size",
						action='store',
						help="Size for vl_dsift features, follow with any number of integer values",
						type=int,
						nargs='*')
	parser.add_argument("--rotation",
						action='store',
						help="degrees of rotation",
						type=int,
						nargs='*')
	parser.add_argument("--num_core",
						help="Number of CPU cores to use in multiprocessing",
						type=int)
	parser.add_argument("--im_size",
					help="Image Height",
					type=int)		
	parser.add_argument("--show_fig",
						help="Show Figure of Misidentified birds",
						type=bool)
	parser.add_argument("--crop",
						help="crop images?",
						type=bool)
	parser.add_argument("--num_words",
						help="Number of centroids found for k-means clustering",
						type=int)
	parser.add_argument("--num_features",
						help="Number of histogram features",
						type=int)

	args = parser.parse_args()

	if args.sample_seed:
		SAMPLE_SEED = args.sample_seed
		if VERBOSE: print ("SAMPLE_SEED = " + str(SAMPLE_SEED))
	seed(SAMPLE_SEED)
	if args.identifier:
		IDENTIFIER = args.identifier
		if VERBOSE: print ("IDENTIFER = " + IDENTIFIER)
	if args.prefix:
		PREFIX = args.prefix
		if VERBOSE: print ("PREFIX = " + PREFIX)

	# Load default configuration
	conf = Configuration(IDENTIFIER, PREFIX)

	# Update configuration from cmd line args
	if args.image_dir:
		conf.calDir = args.image_dir
		if VERBOSE: print ("Image dir: " + conf.calDir)
	if args.num_classes:
		conf.numClasses = args.num_classes
		if VERBOSE: print ("numClasses = " + str(conf.numClasses))
	if args.num_train:
		conf.numTrain = args.num_train
		if VERBOSE: print ("numTrain = " + str(conf.numTrain))
	if args.num_test:
		conf.numTest = args.num_test
		if VERBOSE: print ("numTest = " + str(conf.numTest))
	if args.dsift_size:
		conf.phowOpts.Sizes = args.dsift_size
		if VERBOSE: print ("phowOpts.Sizes = " + str(conf.phowOpts.Sizes))
	if args.rotation:
		conf.rotation = args.rotation
		conf.augment = True
		if VERBOSE: print ("rotation = " + str(conf.rotation))
	if args.num_core:
		conf.numCore = args.num_core
		if VERBOSE: print ("numCore = " + str(conf.numCore))
	if args.im_size:
		conf.imSize = args.im_size
		if VERBOSE: print ("imSize = " + str(conf.imSize))
	if args.num_words:
		conf.numWords = args.num_words
		if VERBOSE: print ("numWords = " + str(conf.numWords))
	if args.num_test or args.num_train:
		conf.imagesperclass = conf.numTest+conf.numTrain
	if args.show_fig:
		conf.showFig = args.show_fig
	if args.crop:
		conf.crop = args.crop
	if args.num_features:
		conf.numbers_of_features_for_histogram = args.num_features
		if VERBOSE: print ("num_features = " + str(conf.numbers_of_features_for_histogram))

	if VERBOSE: print (str(datetime.now()) + ' finished conf')

	classes = get_classes(conf.calDir, conf.numClasses) #get classes from data folder
	print ("Class names" , classes)
	conf.setClasses(classes) #save class names for use by birdid_classifier
	model = Model(classes, conf)

	# all_images_class_labels is an array containing the integer corresponding
	# to the class the image belongs to based on the directory structure
	all_images, all_images_class_labels, images_per_class = get_all_images(classes, conf)
	selTrain, selTest = create_split(all_images, images_per_class, conf)

	if VERBOSE: print (str(datetime.now()) + ' found classes and created split ')

	####################
	# Train vocabulary #
	####################
	if VERBOSE: print (str(datetime.now()) + ' start training vocab')
	if (not exists(conf.vocabPath)) | OVERWRITE:
		vocab = trainVocab(selTrain, all_images, conf)
		print (str(datetime.now()) + ' vocab trained, saving')
		savemat(conf.vocabPath, {'vocab': vocab})
		print (str(datetime.now()) + ' vocab saved')
	else:
		if VERBOSE: print ("using old vocab from " + conf.vocabPath)
		vocab = loadmat(conf.vocabPath)['vocab']
	model.vocab = vocab

	##############################
	# Compute spatial histograms #
	##############################
	if VERBOSE: print (str(datetime.now()) + ' start computing hists')
	if (not exists(conf.histPath)) | OVERWRITE:
		hists = computeHistograms(selTrain, selTest, all_images, model, conf)
		savemat(conf.histPath, {'hists': hists})
	else:
		if VERBOSE: print ("using old hists from " + conf.histPath)
		hists = loadmat(conf.histPath)['hists']

	if conf.augment:
		train = []
		for i in selTrain:
			for x in range(0,len(conf.rotation)+1):
				train.append(i)
		selTrain = train
		test = []
		for i in selTest:
			for x in range(0,len(conf.rotation)+1):
				test.append(i)
		selTest = test

	#######################
	# Compute feature map #
	#######################
	if VERBOSE: print (str(datetime.now()) + ' start computing feature map')
	transformer = AdditiveChi2Sampler()
	histst = transformer.fit_transform(hists)
	train_data = histst[0:len(selTrain)]
	test_data = histst[len(selTrain):]
	
	#############
	# Train SVM #
	#############
	labelsTrain = []
	for t in selTrain:
		labelsTrain.append(all_images_class_labels[t])
	true_classes = []
	for t in selTest:
		true_classes.append(all_images_class_labels[t])
	if (not exists(conf.modelPath)) | OVERWRITE:
		if VERBOSE: print (str(datetime.now()) + ' training liblinear svm')
		if VERBOSE == 'SVM':
			verbose = True
		else:
			verbose = False
		clf = svm.LinearSVC(C=conf.svmC)
		if VERBOSE: print (clf)
		clf.fit(train_data, labelsTrain)
		with open(conf.modelPath, 'wb') as fp:
			dump(clf, fp)
	else:
		if VERBOSE: print ("loading old SVM model")
		with open(conf.modelPath, 'rb') as fp:
			clf = load(fp)

	############
	# Test SVM #
	############
	if (not exists(conf.resultPath)) | OVERWRITE:
		if VERBOSE: print (str(datetime.now()) + ' testing svm')
		predicted_classes = clf.predict(test_data)
		accuracy = accuracy_score(true_classes, predicted_classes)
		cm = confusion_matrix(true_classes, predicted_classes)

		with open(conf.resultPath, 'wb') as fp:
			dump(conf, fp)
			dump(cm, fp)
			dump(predicted_classes, fp)
			dump(true_classes, fp)
			dump(accuracy, fp)
	else:
		with open(conf.resultPath, 'rb') as fp:
			conf = load(fp)
			cm = load(fp)
			predicted_classes = load(fp)
			true_classes = load(fp)
			accuracy = load(fp)

	##################
	# Output Results #
	##################
	print "accuracy =" + str(accuracy)
	newaccuracy = newAccuracy(true_classes, predicted_classes)
	print "new accuracy =" + str(newaccuracy[0])
	print cm
	print str(datetime.now()) + ' run complete with seed = ' + str(SAMPLE_SEED)
	if conf.showFig:
		#Generate Figure of misidentified images
		showFig(newaccuracy[1], conf)
	print str(classification_report(newaccuracy[2], newaccuracy[3], target_names=classes))
	precision = precision_score(newaccuracy[2], newaccuracy[3])
	print str(precision)

	saveCSV("phow_results.xlsx", newaccuracy[0], precision) #save data as excel spreadsheet
