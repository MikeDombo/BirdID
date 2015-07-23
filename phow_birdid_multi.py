#!/usr/bin/env python

"""
Python rewrite of http: //www.vlfeat.org/applications/caltech-101-code.html
"""

from os.path import exists, isdir, basename, join, splitext, isfile
from os import makedirs, remove
from glob import glob
from random import sample, seed
from scipy import ones, mod, arange, array, where, ndarray, hstack, linspace, histogram, vstack, amax, amin
from scipy.misc import imread, imresize
from scipy.cluster.vq import vq
import numpy
from vl_phow import vl_phow
from vlfeat import vl_ikmeans
from scipy.io import loadmat, savemat
from sklearn import svm
from sklearn.metrics import confusion_matrix, accuracy_score
import matplotlib.pyplot as plt
from datetime import datetime
from sklearn.kernel_approximation import AdditiveChi2Sampler
from cPickle import dump, load
import argparse
import multiprocessing
import sys
from openpyxl import *
import ftplib


IDENTIFIER = '2014-04-17-UR'
PREFIX = 'baseline'

FEATUREMAP = True
OVERWRITE = True  # DON'T load mat files generated with a different seed!!!
SAMPLE_SEED = 1963543398
TINYPROBLEM = False
VERBOSE = True	# set to 'SVM' if you want to get the svm output

class Configuration(object):
	def __init__(self, identifier='', prefix=''):
		self.calDir = '../../../data/2014_winter/256x256/vlfeat_training_jpg'

		# Path where training data will be stored
		self.dataDir = '../tempresults'	 # should be resultDir or so
		if not exists(self.dataDir):
			makedirs(self.dataDir)
			print ("folder " + self.dataDir + " created")

		# Sum of these two numbers should be <= # of images in smallest
		# class
		self.numTrain = 70
		self.numTest = 30
		
		self.numCore = multiprocessing.cpu_count()
		self.imagesperclass = self.numTrain + self.numTest
		self.numClasses = 10
		self.numWords = 600
		self.numSpatialX = [2, 4]
		self.numSpatialY = [2, 4]
		self.quantizer = 'vq'  # kdtree from the .m version not implemented
		self.svmC = 15
		self.phowOpts = PHOWOptions(Verbose=False, Sizes=[2, 4, 6, 8], Step=3)
		self.tinyProblem = TINYPROBLEM
		self.prefix = prefix
		self.verbose = True
		self.extensions = [".jpg", ".jpeg", ".bmp", ".png", ".pgm", ".tif", ".tiff"]
		self.images_for_histogram = 30
		self.numbers_of_features_for_histogram = 100000
		self.imSize = 480
		
		self.saveFig = False
		self.showFig = False
		
		self.vocabPath = join(self.dataDir, self.prefix + '-' + identifier + '-vocab.py.mat')
		self.histPath = join(self.dataDir, self.prefix + '-'  + identifier + '-hists.py.mat')
		self.modelPath = join(self.dataDir, self.prefix + '-' + identifier + '-model.py.mat')
		self.resultPath = join(self.dataDir, self.prefix + '-' + identifier + '-result')

		# tests and conversions
		self.phowOpts.Sizes = ensure_type_array(self.phowOpts.Sizes)
		self.numSpatialX = ensure_type_array(self.numSpatialX)
		self.numSpatialY = ensure_type_array(self.numSpatialY)
		if (self.numSpatialX != self.numSpatialY).any():
			messageformat = [str(self.numSpatialX), str(self.numSpatialY)]
			message = "(self.numSpatialX != self.numSpatialY), because {0} != {1}".format(*messageformat)
			raise ValueError(message)

	def setClasses(self, classes):
		self.classes = classes #Save class names for later use by birdid_classifier

def ensure_type_array(data):
	if (type(data) is not ndarray):
		if (type(data) is list):
			data = array(data)
		else:
			data = array([data])
	return data


def standardizeImage(im): #Scales image down to 640x480 or whatever the correct aspect ratio is with conf.imSize as the height
	im = array(im, 'float32') 
	if im.shape[0] > conf.imSize:
		resize_factor = float(conf.imSize) / im.shape[0]	 # don't remove trailing .0 to avoid integer devision
		im = imresize(im, resize_factor)
	if amax(im) > 1.1:
		im = im / 255.0
	assert((amax(im) > 0.01) & (amax(im) <= 1))
	assert((amin(im) >= 0.00))
	return im


def getPhowFeatures(imagedata, phowOpts): #extracts features from image
	im = standardizeImage(imagedata) #scale image to 640x480
	frames, descrs = vl_phow(im,
							 verbose=phowOpts.Verbose,
							 sizes=phowOpts.Sizes,
							 step=phowOpts.Step)
	return frames, descrs

def getPhowFeaturesMulti(imagedata, phowOpts, idx): #used in multiprocessing for training vocab
	return [idx, getPhowFeatures(imagedata, phowOpts)[1]]


def getImageDescriptor(model, im, idx): #gets histograms
	im = standardizeImage(im) #scale image to 640x480
	height, width = im.shape[:2]
	numWords = model.vocab.shape[1]
	frames, descrs = getPhowFeatures(im, conf.phowOpts) #extract features
	# quantize appearance
	if model.quantizer == 'vq':
		binsa, _ = vq(descrs.T, model.vocab.T) #slowest function - assigns words from vocab to features in descrs
	elif model.quantizer == 'kdtree':
		raise ValueError('quantizer kdtree not implemented')
	else:
		raise ValueError('quantizer {0} not known or understood'.format(model.quantizer))
	hist = []
	#generate the histogram bins
	for n_spatial_bins_x, n_spatial_bins_y in zip(model.numSpatialX, model.numSpatialX):
		binsx, distsx = vq(frames[0, :], linspace(0, width, n_spatial_bins_x))
		binsy, distsy = vq(frames[1, :], linspace(0, height, n_spatial_bins_y))
		# binsx and binsy list to what spatial bin each feature point belongs to
		if (numpy.any(distsx < 0)) | (numpy.any(distsx > (width/n_spatial_bins_x+0.5))):
			print ("something went wrong")
			import pdb; pdb.set_trace()
		if (numpy.any(distsy < 0)) | (numpy.any(distsy > (height/n_spatial_bins_y+0.5))):
			print ("something went wrong")
			import pdb; pdb.set_trace()
		# combined quantization
		number_of_bins = n_spatial_bins_x * n_spatial_bins_y * numWords
		temp = arange(number_of_bins)
		# update using this: http://stackoverflow.com/questions/15230179/how-to-get-the-linear-index-for-a-numpy-array-sub2ind
		temp = temp.reshape([n_spatial_bins_x, n_spatial_bins_y, numWords])
		bin_comb = temp[binsx, binsy, binsa]
		hist_temp, _ = histogram(bin_comb, bins=range(number_of_bins+1), density=True) #generate histogram
		hist.append(hist_temp)

	hist = hstack(hist)
	hist = array(hist, 'float32') / sum(hist)
	numTot = float(conf.numClasses*(conf.numTrain+conf.numTest))
	sys.stdout.write ("\r"+str(datetime.now())+" Histograms Calculated: "+str(((idx+1)/numTot)*100.0)[:5]+"%") #make progress percentage
	sys.stdout.flush()
	return [idx, hist]


class Model(object):
	def __init__(self, classes, conf, vocab=None):
		self.classes = classes
		self.phowOpts = conf.phowOpts
		self.numSpatialX = conf.numSpatialX
		self.numSpatialY = conf.numSpatialY
		self.quantizer = conf.quantizer
		self.vocab = vocab


class PHOWOptions(object):
	def __init__(self, Verbose, Sizes, Step):
		self.Verbose = Verbose
		self.Sizes = Sizes
		self.Step = Step


def get_classes(datasetpath, numClasses): #find classes in the data folder
	classes_paths = [files
					 for files in glob(datasetpath + "/*")
					 if isdir(files)]
	classes_paths.sort()
	classes = [basename(class_path) for class_path in classes_paths]
	if len(classes) == 0:
	   raise ValueError('no classes found')
	if len(classes) < numClasses:
	   raise ValueError('conf.numClasses is bigger than the number of folders')
	classes = classes[:numClasses]
	return classes


def get_imgfiles(path, extensions): #get images from 1 class folder
	all_files = []
	all_files.extend([join(path, basename(fname))
					 for fname in glob(path + "/*")
					 if splitext(fname)[-1].lower() in extensions])
	return all_files


def get_all_images(classes, conf): #gets all images from all classes
	all_images = []
	all_images_class_labels = []
	images_per_class = [0]
	for i, imageclass in enumerate(classes):
		path = join(conf.calDir, imageclass)
		extensions = conf.extensions
		imgs = get_imgfiles(path, extensions)
		if len(imgs) == 0:
			raise ValueError('no images for class ' + str(imageclass))
		all_images = all_images + imgs
		if len(images_per_class)>0:
			sum = images_per_class[len(images_per_class)-1]
		images_per_class.append(sum+len(imgs))
		class_labels = list(i * ones(len(imgs)))
		all_images_class_labels = all_images_class_labels + class_labels

	all_images_class_labels = array(all_images_class_labels, 'int')
	return all_images, all_images_class_labels, images_per_class


def create_split(all_images, images_per_class, conf): #split files between training and testing
	train_test = []
	for i in range(0, conf.numClasses):
		new_train = sample(all_images[images_per_class[i]:images_per_class[i+1]], conf.imagesperclass)
		train_test = train_test+new_train
	selTrain = []
	for i in range(0, conf.numClasses):
		selTrain = selTrain + sample(train_test[i*(conf.imagesperclass):(i+1)*(conf.imagesperclass)], conf.numTrain)
	selTest = [x for x in train_test if x not in selTrain]
	for i, ii in enumerate(selTrain):
		selTrain[i] = all_images.index(ii)
	for i, ii in enumerate(selTest):
		selTest[i] = all_images.index(ii)
	return selTrain, selTest


def trainVocab(selTrain, all_images, conf):
	selTrainFeats = sample(selTrain, conf.images_for_histogram)
	descrs = []
	#start multiprocessing block
	pool = multiprocessing.Pool(processes=conf.numCore)
	results = [pool.apply_async(getPhowFeaturesMulti, args=(imread(all_images[ii]), conf.phowOpts, i)) for i, ii in enumerate(selTrainFeats)]
	descrs = [p.get() for p in results]
	pool.terminate()
	sorted(descrs)
	for descr in descrs:
		descr.pop(0)
	depict = []
	for descr in descrs:
		depict.append(descr[0])
	descrs = depict
	#end multiprocessing block
	descrs = hstack(descrs)
	n_features = descrs.shape[1]
	sample_indices = sample(arange(n_features), conf.numbers_of_features_for_histogram)
	descrs = descrs[:, sample_indices]
	descrs = array(descrs, 'uint8')
	
	# Quantize the descriptors to get the visual words
	vocab, _ = vl_ikmeans(descrs,
						  K=conf.numWords,
						  verbose=conf.verbose,
						  method='elkan')
	return vocab

def computeHistograms(all_images, selTrain, selTest, model, conf):
	images = numpy.append(selTrain,selTest).tolist()
	imgs = []
	for i in images:
		imgs.append(all_images[i])
	hists = []
	#start multiprocessing block
	pool = multiprocessing.Pool(processes=conf.numCore)
	results = [pool.apply_async(getImageDescriptor, args=(model, imread(imagefname), ii)) for ii, imagefname in enumerate(imgs)]
	hists = [p.get() for p in results]
	pool.terminate()
	sorted(hists)
	for hist in hists:
		hist.pop(0)
	#end multiprocessing block
	hists = vstack(hists)
	print "" #puts in a new line to separate histogram percentage
	return hists

def saveCSV(file, accuracy):
	dat = []
	dat.append(datetime.now())
	dat.append(str(PREFIX))
	dat.append(str(IDENTIFIER))
	dat.append(str(conf.phowOpts.Sizes))
	dat.append(str(SAMPLE_SEED))
	dat.append(str(accuracy))
	dat.append(str(conf.numTrain))
	dat.append(str(conf.numTest))
	dat.append(str(conf.numClasses))
	dat.append(str(conf.calDir))
	dat.append(str(conf.imSize))
	dat.append(str(conf.numWords))
	dat.append(str(conf.numbers_of_features_for_histogram))

	if isfile("phow_results.xlsx"): #create backup spreadsheet in case network is unmounted
		wb = load_workbook("phow_results.xlsx", guess_types=True)
		ws = wb.active
	else:
		wb = Workbook(guess_types=True)
		ws = wb.active
		ws.append(['Time Completed', 'Prefix', 'Identifier', 'Dsift Sizes', 'Sample Seed', 'Accuracy', 'Number of Train', 'Number of Test', 'Number of Classes', 'Image Path', 'Image Resize Height', 'Number of K-Means Centroids', 'Number of Histogram Features'])
	ws.append(dat)
	wb.save("phow_results.xlsx")

	ftp = ftplib.FTP('dombrowskivpn.mynetgear.com', "lbarnett-students", 'lbarnett-studentaccess')#enter server information here
	ftp.set_pasv(False)
	with open("temp.xlsx", 'wb') as f:
		def callback(data):
			f.write(data)
		ftp.retrbinary('RETR '+str(file), callback)
	wb = load_workbook("temp.xlsx", guess_types=True)
	ws = wb.active
	ws.append(dat)
	wb.save("temp.xlsx")
	ftp.storbinary('STOR '+str(file), open('temp.xlsx','r'))
	ftp.close()
	remove('temp.xlsx')

def showFig(images, conf):
	axes = {}
	if len(images)>0:
		if len(images)<=16:
			x = 4
			y = 4
		elif len(images)<=25:
			x = 5
			y = 5
		elif len(images)<=35:
			x = 7
			y = 5
		else:
			x = 7
			y = 5
			while (int(x)*y)<len(images):
				y = y+1
				x = y*(1.75)
		x = int(x)
		fig = plt.figure(figsize=(16,10))
		for i, im in enumerate(images):
			axes[i] = fig.add_subplot(x,y,i+1)
			axes[i].imshow(imread(im[0]))
			axes[i].get_xaxis().set_ticks([])
			axes[i].get_yaxis().set_ticks([])
			axes[i].set_title("Classified as a "+conf.classes[im[1]['predictedclass']]+"\nActually is a "+conf.classes[im[1]['trueclass']])
	
	fig.set_tight_layout(True)
	plt.show()
	#fig.savefig(conf.output_folder+"/figures/"+imageclass+"/figure_"+str(imName)+".png", dpi=75)


################
# Main Program #
################

if __name__ == '__main__':
	#################################
	# Handle command-line arguments #
	#################################
	parser = argparse.ArgumentParser()
	parser.add_argument("--sample_seed",
		help="Seed for choosing training sample", type=int)

	parser.add_argument("--identifier",
		help="Identifier for this data set. Should not contain the character '-'")
	
	parser.add_argument("--prefix",
		help="Tag used to distinguish versions of a data set")

	parser.add_argument("--image_dir",
						help="Path to directory containing images")

	parser.add_argument("--num_classes",
						help="Number of categories in image set",
						type=int)

	parser.add_argument("--num_train",
						help="Number of training images to use from each category",
						type=int)

	parser.add_argument("--num_test",
						help="Number of test images to use from each catetory",
						type=int)

	parser.add_argument("--dsift_size",
						action='store',
						help="Size for vl_dsift features, follow with any number of integer values",
						type=int,
						nargs='*')
	parser.add_argument("--num_core",
						help="Number of CPU cores to use in multiprocessing",
						type=int)

	parser.add_argument("--im_size",
					help="Image Height",
					type=int)
					
	parser.add_argument("--show_fig",
						help="Show Figure of Misidentified birds",
						type=bool)
		
	parser.add_argument("--num_words",
						help="Number of centroids found for k-means clustering",
						type=int)
	
	parser.add_argument("--num_features",
						help="Number of histogram features",
						type=int)
	
	parser.add_argument("--svm_c",
						help="Number of SVMs competing",
						type=int)

	args = parser.parse_args()

	if args.sample_seed:
		SAMPLE_SEED = args.sample_seed
		if VERBOSE: print ("SAMPLE_SEED = " + str(SAMPLE_SEED))
		
	seed(SAMPLE_SEED)

	if args.identifier:
		IDENTIFIER = args.identifier
		if VERBOSE: print ("IDENTIFER = " + IDENTIFIER)

	if args.prefix:
		PREFIX = args.prefix
		if VERBOSE: print ("PREFIX = " + PREFIX)

	# Load default configuration
	conf = Configuration(IDENTIFIER, PREFIX)

	# Update configuration from cmd line args
	if args.image_dir:
		conf.calDir = args.image_dir
		if VERBOSE: print ("Image dir: " + conf.calDir)

	if args.num_classes:
		conf.numClasses = args.num_classes
		if VERBOSE: print ("numClasses = " + str(conf.numClasses))

	if args.num_train:
		conf.numTrain = args.num_train
		if VERBOSE: print ("numTrain = " + str(conf.numTrain))

	if args.num_test:
		conf.numTest = args.num_test
		if VERBOSE: print ("numTest = " + str(conf.numTest))

	if args.dsift_size:
		conf.phowOpts.Sizes = args.dsift_size
		if VERBOSE: print ("phowOpts.Sizes = " + str(conf.phowOpts.Sizes))

	if args.num_core:
		conf.numCore = args.num_core
		if VERBOSE: print ("numCore = " + str(conf.numCore))

	if args.im_size:
		conf.imSize = args.im_size
		if VERBOSE: print ("imSize = " + str(conf.imSize))

	if args.num_words:
		conf.numWords = args.num_words
		if VERBOSE: print ("numWords = " + str(conf.numWords))

	if args.num_test or args.num_train:
		conf.imagesperclass = conf.numTest+conf.numTrain

	if args.show_fig:
		conf.showFig = args.show_fig

	if args.num_features:
		conf.numbers_of_features_for_histogram = args.num_features
		if VERBOSE: print ("num_features = " + str(conf.numbers_of_features_for_histogram))

	if args.svm_c:
		conf.svmC = args.svm_c
		if VERBOSE: print ("svm_c = " + str(args.svm_c))

	if VERBOSE: print (str(datetime.now()) + ' finished conf')

	classes = get_classes(conf.calDir, conf.numClasses) #get classes from data folder
	print ("Class names" , classes)
	conf.setClasses(classes) #save class names for use by birdid_classifier
	model = Model(classes, conf)

	# all_images_class_labels is an array containing the integer corresponding
	# to the class the image belongs to based on the directory structure
	all_images, all_images_class_labels, images_per_class = get_all_images(classes, conf)
	selTrain, selTest = create_split(all_images, images_per_class, conf)

	if VERBOSE: print (str(datetime.now()) + ' found classes and created split ')

	####################
	# Train vocabulary #
	####################
	if VERBOSE: print (str(datetime.now()) + ' start training vocab')
	if (not exists(conf.vocabPath)) | OVERWRITE:
		vocab = trainVocab(selTrain, all_images, conf)
		print (str(datetime.now()) + ' vocab trained, saving')
		savemat(conf.vocabPath, {'vocab': vocab})
		print (str(datetime.now()) + ' vocab saved')
	else:
		if VERBOSE: print ("using old vocab from " + conf.vocabPath)
		vocab = loadmat(conf.vocabPath)['vocab']
	model.vocab = vocab


	##############################
	# Compute spatial histograms #
	##############################
	if VERBOSE: print (str(datetime.now()) + ' start computing hists')
	if (not exists(conf.histPath)) | OVERWRITE:
		hists = computeHistograms(all_images, selTrain, selTest, model, conf)
		savemat(conf.histPath, {'hists': hists})
	else:
		if VERBOSE: print ("using old hists from " + conf.histPath)
		hists = loadmat(conf.histPath)['hists']


	#######################
	# Compute feature map #
	#######################
	if VERBOSE: print (str(datetime.now()) + ' start computing feature map')
	transformer = AdditiveChi2Sampler()
	histst = transformer.fit_transform(hists)
	train_data = histst[0:conf.numTrain*conf.numClasses]
	test_data = histst[conf.numTrain*conf.numClasses:]

	
	#############
	# Train SVM #
	#############
	if (not exists(conf.modelPath)) | OVERWRITE:
		if VERBOSE: print (str(datetime.now()) + ' training liblinear svm')
		if VERBOSE == 'SVM':
			verbose = True
		else:
			verbose = False
		clf = svm.LinearSVC(C=conf.svmC)
		if VERBOSE: print (clf)
		clf.fit(train_data, all_images_class_labels[selTrain])
		with open(conf.modelPath, 'wb') as fp:
			dump(clf, fp)
	else:
		if VERBOSE: print ("loading old SVM model")
		with open(conf.modelPath, 'rb') as fp:
			clf = load(fp)

	############
	# Test SVM #
	############
	if (not exists(conf.resultPath)) | OVERWRITE:
		if VERBOSE: print (str(datetime.now()) + ' testing svm')
		predicted_classes = clf.predict(test_data)
		true_classes = all_images_class_labels[selTest]
                if VERBOSE:
                        print("True classes", true_classes)
                        print("Pred classes", predicted_classes)
		accuracy = accuracy_score(true_classes, predicted_classes)
		cm = confusion_matrix(true_classes, predicted_classes)

		with open(conf.resultPath, 'wb') as fp:
			dump(conf, fp)
			dump(cm, fp)
			dump(predicted_classes, fp)
			dump(true_classes, fp)
			dump(accuracy, fp)
	else:
		with open(conf.resultPath, 'rb') as fp:
			conf = load(fp)
			cm = load(fp)
			predicted_classes = load(fp)
			true_classes = load(fp)
			accuracy = load(fp)

	##################
	# Output Results #
	##################
	print "accuracy =" + str(accuracy)
	print cm
	print str(datetime.now()) + ' run complete with seed = ' + str(SAMPLE_SEED)
	if conf.showFig:
		#Generate Figure of misidentified images
		misid = []
		for i in range(0, conf.numTest*conf.numClasses):
			if(true_classes[i] != predicted_classes[i]):
				misid.append([all_images[selTest[i]],{'trueclass':true_classes[i],'predictedclass':predicted_classes[i]}])
		showFig(misid, conf)
	saveCSV("phow_results.xlsx", accuracy) #save data as excel spreadsheet
